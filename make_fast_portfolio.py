import yfinance as yf
import pandas as pd
from collections import defaultdict
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import time
from prophet import Prophet
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import os
import win32com.client
import pickle
from datetime import datetime






CACHE_FILE_TEMPLATE = "stock_data_cache_{}.pkl"

def get_cache_file():
    """Generates the cache file name based on the current date."""
    today = datetime.now().strftime("%Y-%m-%d")
    return CACHE_FILE_TEMPLATE.format(today)


def load_cache():
    """Loads the cache from the pickle file."""
    cache_file = get_cache_file()
    if os.path.exists(cache_file):
        with open(cache_file, 'rb') as file:
            return pickle.load(file)
    return {}

def save_cache(cache):
    """Saves the cache to the pickle file."""
    cache_file = get_cache_file()
    with open(cache_file, 'wb') as file:
        pickle.dump(cache, file)



def load_config(file_path):
    with open(file_path, 'r') as file:
        config = json.load(file)
    return config

def get_sp500_tickers():
    url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
    response = requests.get(url)
    tables = pd.read_html(response.text)
    sp500_table = tables[0]
    non_class_c_shares = sp500_table[~sp500_table['Security'].str.contains('Class C', na=False)]
    # Preserve Voting rights
    tickers = non_class_c_shares['Symbol'].tolist()
    return tickers, non_class_c_shares

def categorize_by_sector(sp500_table):
    sector_dict = defaultdict(list)
    for _, row in sp500_table.iterrows():
        sector = row.get('GICS Sector')
        # Skip 'N/A', missing, or empty sectors
        if pd.isna(sector) or sector.strip().upper() in ['N/A', '']:
            continue

        sector_dict[row['GICS Sector']].append(row['Symbol'])
    return sector_dict

def get_sector_performance(sector_dict, period, mode="lazy"):
    #print(sector_dict)
    #print(period)

    cache = load_cache()
    cache_key = f"sector_performance_{period}"
    # If in lazy mode and cached sector data exists, return cached data
    if mode == "lazy" and cache_key in cache:
        print("Loading sector performance from cache.")
        return pd.Series(cache[cache_key], name="Performance")

    sector_performance = {}
    sector_data_cache = {}

    for sector, tickers in sector_dict.items():
        tickers_str = " ".join(tickers)
        try:
            sector_data = yf.download(tickers_str, period=period, group_by='ticker', auto_adjust=True)
            if sector_data.empty:  # Handle empty downloads
                print(f"No data for sector {sector}. Skipping.")
                continue
            sector_close = sector_data.xs('Close', axis=1, level=1)
            if sector_close.empty:  # Handle missing 'Close' data
                print(f"No 'Close' data for sector {sector}. Skipping.")
                continue
              # Cache sector data
            sector_data_cache[sector] = sector_data
            sector_performance[sector] = (sector_close.iloc[-1].sum() / sector_close.iloc[0].sum()) - 1
        except Exception as e:
            print(f"Error fetching sector performance for {sector}: {e}")
            sector_performance[sector] = None  # Mark as failed

    cache[f"sector_data_{period}"] = sector_data_cache
    cache[cache_key] = sector_performance
    save_cache(cache)

    return pd.Series(sector_performance, name="Performance")


def analyze_fundamentals(stock_ticker,sector,mode='lazy'):

    cache = load_cache()
    ticker_key = f"fundamentals_{stock_ticker}"

    if mode == "lazy" and ticker_key in cache:
        print(f"Loading cached fundamentals for {stock_ticker}.")
        return cache[ticker_key]
    
    time.sleep(1)
    stock = yf.Ticker(stock_ticker)
    try:
            
        info = stock.info
        
        eps = info.get("trailingEps")
        revenue = info.get("totalRevenue")
        market_cap = info.get("marketCap")
        pe_ratio = info.get("trailingPE")
        price_to_sales = info.get("priceToSalesTrailing12Months")
        price_to_book = info.get("priceToBook")
        ebitda = info.get("ebitda")
        profit_margins = info.get("profitMargins")
        share_price = info.get("currentPrice")


        fundamentals = {
            "Ticker": stock_ticker,
            "Sector": sector,
            "EPS": info.get("trailingEps"),
            "Revenue": info.get("totalRevenue"),
            "MarketCap": info.get("marketCap"),
            "PE_Ratio": info.get("trailingPE"),
            "Price_to_Sales": info.get("priceToSalesTrailing12Months"),
            "Price_to_Book": info.get("priceToBook"),
            "EBITDA": info.get("ebitda"),
            "Profit_Margins": info.get("profitMargins"),
            "Share_Price": info.get("currentPrice"),
        }


        cache[ticker_key] = fundamentals
        save_cache(cache) 
        return fundamentals
    except Exception as e:
        print(f"Error fetching data for {stock_ticker}: {e}")
        return None

# USD Stocks We must Convert To Selected Currancy Based On Excange Rate 
def get_exchange_rate(currency):
        currency = str(currency).upper()
        url = 'https://api.exchangerate-api.com/v4/latest/USD'
        try:
            response = requests.get(url)
            response.raise_for_status()  # Raise HTTPError for bad responses
            data = response.json()
        except requests.RequestException as e:
            raise Exception(f"Failed to fetch exchange rates: {e}")
        
        rates = data.get("rates", {})
        if currency in rates:
            return rates[currency]
        else:
            raise Exception(f"Currency '{currency}' not found in exchange rates : {str(rates.keys())}")
        



def _get_dynamic_picks(config, sector_dict, remaining_percentage, existing_picks=0,excluded_tickers=None):
    """
    Helper function to dynamically select stocks for the remaining percentage.
    """
    print("Selecting dynamic picks to fill the gap.")
    sector_performance = get_sector_performance(
        sector_dict,
        config["performance_period"],
        mode=config.get("cache", "lazy")
    )
    print("Sector Performance (YTD):\n", sector_performance)

    banned_sectors = config.get("banned_sectors", [])
    top_n_sectors = config.get("top_n_sectors", 3)

    top_sectors = [
        sector
        for sector in sector_performance.nlargest(len(sector_performance)).index
        if sector not in banned_sectors
    ][:top_n_sectors]
    print("Top Sectors:", top_sectors)

    stock_candidates = []
    for sector in top_sectors:
        stocks = sector_dict[sector]
        print(f"Stocks in sector {sector}: {stocks}")

        for stock in stocks:
            if stock in excluded_tickers:
                continue
            fundamentals = analyze_fundamentals(
                stock,
                sector,
                mode=config.get("cache", "lazy")
            )
            if fundamentals:
                stock_candidates.append(fundamentals)

    # Create a DataFrame and select the top stocks
    df = pd.DataFrame(stock_candidates)
    df = df.dropna()

    # Filter and rank the stocks
    df = df[
        (df["EPS"] > 0) & (df["Revenue"] > 0) & (df["MarketCap"] > 0) & (df["PE_Ratio"] > 0)
    ]
    if config["filter_pe_ratio"]:
        df = df[df["PE_Ratio"] < config["pe_ratio_threshold"]]

    df["Score"] = (
        df["EPS"].rank(ascending=False) +
        df["Revenue"].rank(ascending=False) +
        df["MarketCap"].rank(ascending=False)
    )
    df = df.sort_values(by="Score", ascending=True)

    # Determine the number of additional picks required
    num_additional_picks = config["num_picks"] - existing_picks
    dynamic_picks = df.head(num_additional_picks).to_dict(orient="records")

    # Assign remaining percentage to each dynamic pick
    remaining_percentage_per_stock = remaining_percentage / len(dynamic_picks)
    for pick in dynamic_picks:
        pick["Percentage"] = remaining_percentage_per_stock

    return dynamic_picks, sector_performance



def select_top_stocks(config, **kwargs):
    """
    Selects top stocks based on configuration and optional stock picks.
    Handles:
    - Stock picks as a dictionary with percentages.
    - Dynamic selection to fill missing percentages or picks.
    """
    stock_picks = kwargs.get("stock_picks", None)
    stock_candidates = []
    sector_performance = None 

    # Fetch tickers and sector data
    tickers, sp500_table = get_sp500_tickers()
    sector_dict = categorize_by_sector(sp500_table)

    sector_performance = get_sector_performance(
        sector_dict,
        config["performance_period"],
        mode=config.get("cache", "lazy")
    )
    print("Sector Performance (YTD):\n", sector_performance)

    # Handle provided stock picks as a dictionary
    if isinstance(stock_picks, dict):
        print("Using provided stock picks as a dictionary with percentages.")
        total_percentage = sum(stock_picks.values())

        # Normalize percentages if they exceed 100%
        if total_percentage > 100:
            print(f"Adjusting percentages to fit 100% (original total: {total_percentage}%).")
            stock_picks = {k: v / total_percentage * 100 for k, v in stock_picks.items()}

        # Process fundamentals for provided stock picks
        for stock, percentage in stock_picks.items():
            fundamentals = analyze_fundamentals(
                stock,
                "N/A",  # Sector is unknown for provided picks
                mode=config.get("cache", "lazy")
            )
            if fundamentals:
                fundamentals["Percentage"] = percentage
                stock_candidates.append(fundamentals)

        # Handle remaining percentage if under 100%
        if total_percentage < 100:
            remaining_percentage = 100 - total_percentage
            print(f"Filling the remaining {remaining_percentage}% with dynamic picks.")


            # Dynamically select stocks
            dynamic_stocks, _ = _get_dynamic_picks(
                config,
                sector_dict,
                remaining_percentage,
                len(stock_picks),
                excluded_tickers=set(stock_picks.keys())
            )
            stock_candidates.extend(dynamic_stocks)

    # Handle case when no stock picks are provided
    else:
        print("No stock picks provided. Selecting dynamically based on sector performance.")
        dynamic_stocks, sector_performance = _get_dynamic_picks(
            config,
            sector_dict,
            100  # Full allocation for dynamic picks
        )
        stock_candidates.extend(dynamic_stocks)

    # Create a DataFrame from the final stock candidates
    df = pd.DataFrame(stock_candidates)
    print("DataFrame before cleaning:\n", df)
    #Re Fetch Kwargs 
    provided_tickers = set(kwargs.get("stock_picks", {}).keys()) if isinstance(stock_picks, dict) else set()
    #Clean
    df = df[ df["Ticker"].isin(provided_tickers) | (~df.isna().any(axis=1))]

    # Convert numeric columns
    numeric_columns = [
        "EPS", "Revenue", "MarketCap", "PE_Ratio",
        "Price_to_Sales", "Price_to_Book", "EBITDA", "Profit_Margins", "Share_Price"
    ]
    df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors="coerce")

    #Clean
    df = df[ df["Ticker"].isin(provided_tickers) | (~df.isna().any(axis=1))]

    #Re Fetch Kwargs 
    # Apply filters
    df = df[
        ((df["EPS"] > 0) & (df["Revenue"] > 0) & (df["MarketCap"] > 0) & (df["PE_Ratio"] > 0))| df["Ticker"].isin(provided_tickers) 
    ]

    if config["filter_pe_ratio"]:
        df = df[(df["PE_Ratio"] < config["pe_ratio_threshold"] )| df["Ticker"].isin(provided_tickers) ]

    print("DataFrame after filtering positive values and PE Ratio:\n", df)

    # Rank and normalize the values
    df["EPS_Rank"] = df["EPS"].rank(ascending=False)
    df["Revenue_Rank"] = df["Revenue"].rank(ascending=False)
    df["MarketCap_Rank"] = df["MarketCap"].rank(ascending=False)

    # Normalize the ranks
    df["EPS_Score"] = df["EPS_Rank"] / df["EPS_Rank"].max()
    df["Revenue_Score"] = df["Revenue_Rank"] / df["Revenue_Rank"].max()
    df["MarketCap_Score"] = df["MarketCap_Rank"] / df["MarketCap_Rank"].max()

    # Calculate the final score
    df["Score"] = df["EPS_Score"] + df["Revenue_Score"] + df["MarketCap_Score"]

    df = df.dropna(subset=["Score"])
    print("DataFrame after calculating Score:\n", df)

    # Select top stocks
    top_stocks = df.nsmallest(config["num_picks"], "Score")

    print("Sector Performance  Top Stocks END (YTD):\n", sector_performance)

    return top_stocks, sector_performance




def forecast_with_prophet(data, periods):
    df = data.reset_index()
    df.columns = ['ds', 'y']
    df['ds'] = df['ds'].dt.tz_localize(None)  # Remove timezone information
    model = Prophet()
    model.fit(df)
    future = model.make_future_dataframe(periods=periods)
    forecast = model.predict(future)
    return forecast[['ds', 'yhat']]


# Builds portfolio for you based on top stocks Data_frame picks
def build_portfolio(top_stocks, config):
    crncy = config["currency"]
    exchange_rate = get_exchange_rate(currency=crncy)
    flat_fee = config["flat_fee"]
    total_value = config["total_value"]

    # Convert share prices to CAD
    crncy_identifier =f'Share_Price_{crncy}'
    top_stocks[crncy_identifier] = top_stocks['Share_Price'] * exchange_rate

    if "Percentage" in top_stocks.columns:
        top_stocks["Investment"] = (top_stocks["Percentage"] / 100) * total_value
    else:
         top_stocks["Investment"] = total_value / len(top_stocks)



    # Calculate the initial evenly distributed investment amount
    #initial_investment_per_stock = total_value / len(top_stocks)

    # Initialize the number of shares to buy for each stock
    top_stocks['Num_Shares'] = (top_stocks["Investment"]  / (top_stocks[crncy_identifier] + flat_fee)).apply(lambda x: int(x))

    # Calculate the initial investment for each stock
    top_stocks['Investment'] = (top_stocks['Num_Shares'] * top_stocks[crncy_identifier]) + flat_fee

    # Calculate the remaining value to invest
    total_invested = top_stocks['Investment'].sum()
    remaining_value = total_value - total_invested

    # Adjust for the remainder to distribute any leftover funds
    while remaining_value >= top_stocks[crncy_identifier].min():
        for index, row in top_stocks.iterrows():
            additional_investment = row[crncy_identifier] 
            if remaining_value >= additional_investment:
                top_stocks.at[index, 'Num_Shares'] += 1
                top_stocks.at[index, 'Investment'] += additional_investment
                remaining_value -= additional_investment

    return top_stocks


def add_to_excel(top_stocks, sector_performance, filename):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top Stocks"
    
    for r in dataframe_to_rows(top_stocks, index=False, header=True):
        ws1.append(r)
    
    ws2 = wb.create_sheet(title="Sector Performance")
    sector_df = pd.DataFrame(sector_performance, columns=['Performance'])
    for r in dataframe_to_rows(sector_df.reset_index(), index=False, header=True):
        ws2.append(r)
    
    wb.save(filename)

def fetch_stock_forecast(top_stocks, t_period, mode="lazy"):
    """
    Fetches stock data for forecasts, utilizing caching.

    Args:
        top_stocks (pd.DataFrame): DataFrame of top stocks with a "Ticker" column.
        t_period (str): Period for fetching stock data.
        mode (str): Mode of operation ("lazy" to use cache, fetch otherwise).

    Returns:
        dict: Dictionary with tickers as keys and stock data as values.
    """
    cache = load_cache()
    forecasts = {}

    for _, stock in top_stocks.iterrows():
        ticker = stock["Ticker"]
        ticker_key = f"forecast_data_{ticker}_{t_period}"

        if mode == "lazy" and ticker_key in cache:
            print(f"Loading cached data for {ticker}.")
            forecasts[ticker] = cache[ticker_key]
        else:
            try:
                print(f"Downloading data for {ticker}.")
                stock_data = yf.download(ticker, period=t_period, auto_adjust=True)["Close"]
                forecasts[ticker] = stock_data
                cache[ticker_key] = stock_data
            except Exception as e:
                print(f"Error fetching data for {ticker}: {e}")

    save_cache(cache)
    return forecasts


def save_forecast_to_excel_predict(top_stocks, sector_performance, filename,config):
        def close_excel_file_if_open(filepath):
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                for workbook in excel.Workbooks:
                    if workbook.FullName == os.path.abspath(filepath):
                        workbook.Close(False)  # Close without saving changes
                        excel.Quit()
                        break
            except Exception as e:
                raise Exception(f"Failed to close Excel file: {e}")

        # Check if the file already exists and handle it
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except PermissionError:
                # Attempt to close the Excel window holding the file
                close_excel_file_if_open(filename)
                # Retry deleting the file after closing it
                try:
                    os.remove(filename)
                except Exception as e:
                    raise Exception(f"Unable to delete the file {filename} after closing Excel: {e}")


        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Top Stocks"
        t_period = config["historical"]
        
        for r in dataframe_to_rows(top_stocks, index=False, header=True):
            ws1.append(r)
        
        ws2 = wb.create_sheet(title="Sector Performance")
        sector_df = pd.DataFrame(sector_performance, columns=['Performance'])
        for r in dataframe_to_rows(sector_df.reset_index(), index=False, header=True):
            ws2.append(r)
        
        # Create a sheet for each stock's forecast

        forcasts  = fetch_stock_forecast(top_stocks, t_period, mode=config.get('cache','lazy'))
        for _, stock in top_stocks.iterrows():
            ticker = stock["Ticker"]
            stock_data = forcasts[ticker]

            #arima_forecast = forecast_with_arima(stock_data, periods=30)
            prophet_forecast1 = forecast_with_prophet(stock_data, periods=180)
            prophet_forecast2 = forecast_with_prophet(stock_data, periods=60)
            prophet_forecast3 = forecast_with_prophet(stock_data, periods=30)
            prophet_forcasts = {"180 day":prophet_forecast1,"60 day" :prophet_forecast2, "30 day": prophet_forecast3  }
            # Generate the forecast visualization
            img_stream = generate_forecast_image(ticker, stock_data,  prophet_forcasts)

            # Add the image to a new sheet
            ws_forecast = wb.create_sheet(title=f"{ticker}_Forecast")
            img = OpenpyxlImage(img_stream)
            ws_forecast.add_image(img, "A1")

            # Calculate and add Prophet percentage changes
            try:
                actual_price = stock_data.iloc[-1].values[0]  # Latest historical price
            except:
                    pass
            try:
                stock_data['Close'].iloc[-1]  # Latest historical price
            except:
                    pass

            ws_forecast.append(["Detail", "Value"])
            ws_forecast.append(["Actual Price", actual_price])

            row_offset = ws_forecast.max_row + 2  # Leave space for visual clarity
            ws_forecast.cell(row=row_offset, column=1, value="Forecast Period")
            ws_forecast.cell(row=row_offset, column=2, value="Forecasted Price")
            ws_forecast.cell(row=row_offset, column=3, value="Percent Change")
            ws_forecast.cell(row=row_offset, column=4, value="Prophet Percentage")
            row_offset += 1

            for label, forecast in prophet_forcasts.items():
                forecast_price = forecast['yhat'].iloc[-1]  # Last forecasted price
                percent_change = (forecast_price - actual_price) / actual_price * 100
                prophet_percent = percent_change  # Could add more logic here if needed
                
                # Write details into Excel
                ws_forecast.cell(row=row_offset, column=1, value=label)
                ws_forecast.cell(row=row_offset, column=2, value=forecast_price)
                ws_forecast.cell(row=row_offset, column=3, value=f"{percent_change:.2f}%")
                ws_forecast.cell(row=row_offset, column=4, value=f"{prophet_percent:.2f}%")
                row_offset += 1



        wb.save(filename)

def generate_forecast_image(ticker, historical, prophet_forecasts):
    #two_years_ago = historical.index[-1] - pd.Timedelta(days=730)
    #historical_filtered = historical[historical.index >= two_years_ago]
    plt.figure(figsize=(12, 6))
    plt.plot(historical, label="Historical Data", color="blue")
    #plt.plot(range(len(historical), len(historical) + len(arima_forecast)), arima_forecast, label="ARIMA Forecast", color="orange")
    forcast = None
    colors = ['green', 'orange', 'purple', 'red', 'cyan']
    for idx, (label, forecast) in enumerate(prophet_forecasts.items()):
        color = colors[idx % len(colors)]  # Cycle through colors if more forecasts than colors
        plt.plot(forecast['ds'], forecast['yhat'], label=f"{label} Forecast", color=color, linewidth=2)
    """for label, forecast in prophet_forecasts.items():
        plt.plot(forecast['ds'], forecast['yhat'], label=f"{label} Forecast", linewidth=2)
    for forecast in prophet_forecasts:
        plt.plot(forecast['ds'], forecast['yhat'], label="Prophet Forecast", color="green")"""
    plt.title(f"Forecast for {ticker}")
    plt.xlabel("Date")
    plt.ylabel("Price")
    plt.legend()
    plt.grid()
    
    # Save the plot to a BytesIO stream
    img_stream = BytesIO()
    plt.savefig(img_stream, format='png', bbox_inches='tight')
    img_stream.seek(0)
    plt.close()
    return img_stream

if __name__ == "__main__":
    TEST_MODE = True
    config = load_config('config.json')
    if TEST_MODE:
        from cgemini1206exp import generate_portfolio_config
        port_request = input("What portfolio would you like?: ")
        print(f"Requested Portfolio with : { port_request}")
        config = generate_portfolio_config(port_request)
    stock_picks = config.get("picks",None)
    if not stock_picks or not isinstance(stock_picks, (dict, list)):
        top_stocks, sector_performance = select_top_stocks(config)
        print("Top Stocks:\n", top_stocks)
    elif isinstance(stock_picks, dict):
        top_stocks, sector_performance = select_top_stocks(config, stock_picks=stock_picks)
    elif isinstance(stock_picks, list):
        print("Using provided stock list.")
        stock_picks = stock_picks[:config.get("num_picks", len(stock_picks))]
        equal_percentage = 100 / len(stock_picks)
        stock_picks = {ticker: equal_percentage for ticker in stock_picks}
        top_stocks, sector_performance = select_top_stocks(config, stock_picks=stock_picks)
    else:
        top_stocks, sector_performance = select_top_stocks(config)
        print("Top Stocks:\n", top_stocks)
    
    top_stocks = build_portfolio(top_stocks, config)
    print("Portfolio:\n", top_stocks)
    
    
    output_file =  config.get("output_filename","portfolio.xlsx")
    save_forecast_to_excel_predict(top_stocks, sector_performance, output_file,config)
    print(f"Data saved to '{output_file}'")
